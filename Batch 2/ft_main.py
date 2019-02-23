from openpyxl import load_workbook

class user:

    def Inrate():
        pass

    def SGA():
        pass

class insert:

    def newFeedback():
        print("**Please insert below information**")
        CompanyID = input("Company ID = ")
        InrateAnalyst = input("Inrate analyst name = ")
        SGAAnalyst  = input("SGA analyst name  = ")
        FeedbackType = input("Feedback type = ")
        Indicator = input("Indicator name = ")
        InrateRemark = input("Inrate comment = ")
        newRow =(CompanyID,InrateAnalyst,SGAAnalyst,FeedbackType,Indicator,InrateRemark,SGARemark)
        wb = load_workbook(filename='ft_file.xlsx')
        ws = wb['Sheet1']
        ws.append(newRow)
        print("Saved successfully!")
        wb.save("ft_file.xlsx")

    def editFeedback():
        check.all()
        print("**Which feedback you want to edit again?**")
        checkID = input("Mention company ID = ")
        wb = load_workbook(filename='ft_file.xlsx')
        ws = wb['Sheet1']

        for row in ws.rows:
            if row[0].value == checkID:
                row_edit = row
                print(row_edit)
                break

        for r in row_edit:
            print(r.value)
            i = input("Do you want to change this value? (y/n) = ")
            if i == "y":
                r.value = input("Enter new value = ")

        print("Saved successfully!")
        wb.save("ft_file.xlsx")

    def response():
        check.all()
        print("**Which feedback you want to respond to?**")
        checkID = input("Mention company ID = ")
        wb = load_workbook(filename='ft_file.xlsx')
        ws = wb['Sheet1']

        for row in ws.rows:
            if row[0].value == checkID:
                row_edit = row
                print(row_edit)
                break

        for r in row_edit:
            print(r.value)
            i = input("Do you want to change this value? (y/n) = ")
            if i == "y":
                r.value = input("Enter new value = ")

        print("Saved successfully!")
        wb.save("ft_file.xlsx")


class check:

    def all():
        wb = load_workbook(filename='ft_file.xlsx', read_only=True)
        ws = wb['Sheet1']

        for row in ws.rows:
            row_text =""
            for x in row:
                row_text = row_text + str(x.value) + "||"
            print(row_text)
            # for cell in row:
            #     print(cell.value)


    def byUser():
        pass

    def byStatus():
        pass


if __name__ == "__main__":


    #insert.newFeedback()
    insert.response()
    check.all()
    #insert.editFeedback()
