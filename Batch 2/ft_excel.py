from openpyxl import load_workbook

class User:

    def __init__(self):
        pass

    def exrate(self):
        pass

    def ags(self):
        pass

class Insert:

    def __init__(self):
        pass

    def newFeedback(self):
        print("**Please insert below information**")
        CompanyID = input("Company ID = ")
        ExrateAnalyst = input("Exrate analyst name = ")
        AGSAnalyst  = input("AGS analyst name  = ")
        FeedbackType = input("Feedback type = ")
        Indicator = input("Indicator name = ")
        ExrateRemark = input("Exrate comment = ")
        # AGSRemark=""
        # Status=""
        newRow =(CompanyID,ExrateAnalyst,AGSAnalyst,FeedbackType,Indicator,ExrateRemark)
        wb = load_workbook(filename='ft_file.xlsx')
        ws = wb['Sheet1']
        ws.append(newRow)
        print("Feedback inserted successfully!")
        wb.save("ft_file.xlsx")

    def editFeedback(self):
        self.all()
        print("**Which feedback you want to edit again?**")
        checkID = input("Mention company ID = ")
        wb = load_workbook(filename='ft_file.xlsx')
        ws = wb['Sheet1']

        for row in ws.rows:
            if row[0].value == checkID:
                row_edit = row
                print(row_edit)
                break

        for r in row_edit:
            print(r.value)
            i = input("Do you want to change this value? (y/n) = ")
            if i == "y":
                r.value = input("Enter new value = ")

        print("Edited successfully!")
        wb.save("ft_file.xlsx")

    def response(self):
        self.all()
        print("**Which feedback you want to respond to?**")
        checkID = input("Mention company ID = ")
        wb = load_workbook(filename='ft_file.xlsx')
        ws = wb['Sheet1']

        for row in ws.rows:
            if row[0].value == checkID:
                row_edit = row
                #print(row_edit)
                break

        for r in row_edit:
            print(r.value)
            i = input("Do you want to change this value? (y/n) = ")
            if i == "y":
                r.value = input("Enter new value = ")

        print("Response inserted successfully!")
        wb.save("ft_file.xlsx")


class Check(Insert):

    def __init__(self):
        pass

    def all(self):
        wb = load_workbook(filename='ft_file.xlsx', read_only=True)
        ws = wb['Sheet1']

        for row in ws.rows:
            row_text =""
            for x in row:
                row_text = row_text + str(x.value) + "||"
            print(row_text)
            # for cell in row:
            #     print(cell.value)

    def welcomeScreen(self):
        print("----Welcome to AGS Exrate Feedback Tool----")
        print("[1] Insert feedback (for Exrate analysts)")
        print("[2] Insert response (for AGS analysts)")
        print("[3] Edit feedback (for Exrate analysts)")
        print("[4] Edit response (for AGS analysts)")
        print("[5] Check FT worksheet")
        print("[0] Exit")
        infut = input("Please enter your input number = ")
        if infut== "1":
            self.newFeedback()
        elif infut == "2":
            self.response()
        elif infut == "3":
            self.editFeedback()
        elif infut == "4":
            self.editFeedback()
        elif infut == "5":
            self.all()
        elif infut == "0":
            pass

    def byUser(self):
        pass

    def byStatus(self):
        pass


if __name__ == "__main__":

    ft = Check()
    ft.welcomeScreen()
